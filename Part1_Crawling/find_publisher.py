import requests
import openpyxl
from concurrent.futures import ThreadPoolExecutor

def get_data(data,sheet_2):
    headers={
        'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'
    }
    url='https://data.europa.eu/api/hub/search/datasets/'+data
    resp=requests.get(url,headers=headers)
    text=resp.json()
    try:
        publisher=text['result']['publisher']['name']
    except:
        publisher=''
    sheet_2.append([data,publisher])

if __name__ == '__main__':

    wd_2=openpyxl.Workbook()
    wd=openpyxl.load_workbook('data.xlsx')
    names=wd.sheetnames
    names.reverse()
    for name in names:
        print(name)
        wd_2.create_sheet(name)
        sheet_2=wd_2[wd_2.sheetnames[-1]]
        sheet=wd[name]
        with ThreadPoolExecutor(50) as t:
            for i in range(1,sheet.max_row+1):
                t.submit(get_data,sheet.cell(row=i,column=1).value,sheet_2)
        wd_2.save('publisher.xlsx')

