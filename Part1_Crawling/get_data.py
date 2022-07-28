import requests
import openpyxl
wd=openpyxl.load_workbook('data.xlsx')
wd.create_sheet('INTR')
sheet=wd[wd.sheetnames[-1]]
headers={
    'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'
}

data={
'q': '',
'filter': 'dataset',
'limit': '1000',
'page': '0',
'sort': 'relevance+desc, modified+desc, title.en+asc',
'facetOperator': 'AND',
'facetGroupOperator': 'AND',
'dataServices': 'false',
'includes': 'id,title.en,description.en,languages,modified,issued,catalog.id,catalog.title,catalog.country.id,distributions.id,distributions.format.label,distributions.format.id',
'facets': '{"country":[],"catalog":[],"categories":["INTR"],"publisher":[],"keywords":[],"dataServices":[],"scoring":[],"format":[],"license":[]}'
}
for i in range(0,4):
    print(i)
    url='https://data.europa.eu/api/hub/search/search'
    data['page']=f'{i}'
    resp=requests.get(url,headers=headers,params=data)
    texts=resp.json()['result']['results']
    for text in texts:
        try:
            title=text['title']['en']
            url='https://data.europa.eu/data/datasets/'+text['id']+'?locale=en'
            catalog=text['catalog']['id']
            sheet.append([url,title,catalog])
        except:
            print('error')
            continue
    wd.save('data.xlsx')

